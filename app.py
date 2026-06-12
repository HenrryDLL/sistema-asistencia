from flask import Flask, render_template, request, redirect, url_for, session
from datetime import datetime
import pytz

app = Flask(__name__)
app.secret_key = 'liarfiresystem2024'

def hora_peru():
    zona_peru = pytz.timezone('America/Lima')
    ahora = datetime.now(zona_peru)
    return ahora.strftime('%Y-%m-%d'), ahora.strftime('%H:%M:%S')

@app.route('/')
def inicio():
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    usuario = request.form['usuario']
    contrasena = request.form['contrasena']

    import sqlite3
    conexion = sqlite3.connect('asistencia.db')
    cursor = conexion.cursor()
    cursor.execute('SELECT * FROM usuarios WHERE usuario = ? AND contrasena = ?', 
                   (usuario, contrasena))
    user = cursor.fetchone()
    conexion.close()

    if user:
        session['usuario'] = usuario
        return redirect(url_for('dashboard'))
    else:
        return render_template('login.html', error='Usuario o contraseña incorrectos')

@app.route('/dashboard')
def dashboard():
    if 'usuario' not in session:
        return redirect(url_for('inicio'))
    return render_template('dashboard.html', usuario=session['usuario'])

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('inicio'))

@app.route('/trabajadores')
def trabajadores():
    if 'usuario' not in session:
        return redirect(url_for('inicio'))
    
    import sqlite3
    conexion = sqlite3.connect('asistencia.db')
    cursor = conexion.cursor()
    cursor.execute('SELECT id, nombre, dni, obra, cargo FROM trabajadores')
    lista = cursor.fetchall()
    conexion.close()
    
    return render_template('trabajadores.html', 
                           usuario=session['usuario'], 
                           trabajadores=lista)

@app.route('/trabajadores/registrar', methods=['POST'])
def registrar_trabajador():
    if 'usuario' not in session:
        return redirect(url_for('inicio'))

    nombre = request.form['nombre']
    dni = request.form['dni']
    cargo = request.form['cargo']
    obra = request.form['obra']
    
    import sqlite3
    conexion = sqlite3.connect('asistencia.db')
    cursor = conexion.cursor()
    cursor.execute('INSERT INTO trabajadores (nombre, dni, cargo, obra) VALUES (?, ?, ?, ?)',
                   (nombre, dni, cargo, obra))
    conexion.commit()
    conexion.close()
    
    return redirect(url_for('trabajadores'))

@app.route('/asistencia')
def asistencia():
    if 'usuario' not in session:
        return redirect(url_for('inicio'))
    
    import sqlite3
    conexion = sqlite3.connect('asistencia.db')
    conexion.row_factory = sqlite3.Row
    cursor = conexion.cursor()
    
    hoy, _ = hora_peru()
    
    cursor.execute('''
        SELECT t.id, t.nombre, t.dni, t.cargo, t.obra,
               a.hora_ingreso, a.hora_salida
        FROM trabajadores t
        LEFT JOIN asistencia a 
        ON t.id = a.trabajador_id AND a.fecha = ?
    ''', (hoy,))
    
    trabajadores = cursor.fetchall()
    conexion.close()
    
    return render_template('asistencia.html', 
                           usuario=session['usuario'],
                           trabajadores=trabajadores)

@app.route('/asistencia/entrada', methods=['POST'])
def registrar_entrada():
    if 'usuario' not in session:
        return redirect(url_for('inicio'))
    
    import sqlite3
    
    trabajador_id = request.form['trabajador_id']
    # NUEVO: recibe foto y ubicación del formulario
    foto = request.form.get('foto', '')
    ubicacion = request.form.get('ubicacion', '')
    hoy, hora = hora_peru()
    
    conexion = sqlite3.connect('asistencia.db')
    cursor = conexion.cursor()
    cursor.execute('''
        INSERT INTO asistencia (trabajador_id, hora_ingreso, fecha, foto_ingreso, ubicacion_ingreso)
        VALUES (?, ?, ?, ?, ?)
    ''', (trabajador_id, hora, hoy, foto, ubicacion))
    conexion.commit()
    conexion.close()
    
    return redirect(url_for('asistencia'))

@app.route('/asistencia/salida', methods=['POST'])
def registrar_salida():
    if 'usuario' not in session:
        return redirect(url_for('inicio'))
    
    import sqlite3
    
    trabajador_id = request.form['trabajador_id']
    # NUEVO: recibe foto y ubicación del formulario
    foto = request.form.get('foto', '')
    ubicacion = request.form.get('ubicacion', '')
    hoy, hora = hora_peru()
    
    conexion = sqlite3.connect('asistencia.db')
    cursor = conexion.cursor()
    cursor.execute('''
        UPDATE asistencia SET hora_salida = ?, foto_salida = ?, ubicacion_salida = ?
        WHERE trabajador_id = ? AND fecha = ?
    ''', (hora, foto, ubicacion, trabajador_id, hoy))
    conexion.commit()
    conexion.close()
    
    return redirect(url_for('asistencia'))

@app.route('/reportes')
def reportes():
    if 'usuario' not in session:
        return redirect(url_for('inicio'))
    
    import sqlite3
    
    hoy, _ = hora_peru()
    fecha_seleccionada = request.args.get('fecha', hoy)
    
    conexion = sqlite3.connect('asistencia.db')
    conexion.row_factory = sqlite3.Row
    cursor = conexion.cursor()
    
    # NUEVO: incluye foto y ubicación en la consulta
    cursor.execute('''
        SELECT t.nombre, t.dni, t.cargo, t.obra,
               a.hora_ingreso, a.hora_salida,
               a.foto_ingreso, a.ubicacion_ingreso,
               a.foto_salida, a.ubicacion_salida
        FROM trabajadores t
        LEFT JOIN asistencia a 
        ON t.id = a.trabajador_id AND a.fecha = ?
    ''', (fecha_seleccionada,))
    
    reportes = cursor.fetchall()
    conexion.close()
    
    return render_template('reportes.html',
                           usuario=session['usuario'],
                           reportes=reportes,
                           fecha_seleccionada=fecha_seleccionada)

@app.route('/produccion')
def produccion():
    if 'usuario' not in session:
        return redirect(url_for('inicio'))
    
    import sqlite3
    
    fecha_inicio = request.args.get('fecha_inicio', '')
    fecha_fin = request.args.get('fecha_fin', '')
    
    conexion = sqlite3.connect('asistencia.db')
    conexion.row_factory = sqlite3.Row
    cursor = conexion.cursor()
    
    query = '''
        SELECT t.nombre, t.cargo, t.obra,
               COUNT(a.fecha) as dias_elaborados,
               SUM(
                   CASE 
                       WHEN a.hora_ingreso IS NOT NULL AND a.hora_salida IS NOT NULL
                       THEN ROUND((julianday(a.hora_salida) - julianday(a.hora_ingreso)) * 24, 2)
                       ELSE 0
                   END
               ) as total_horas
        FROM trabajadores t
        LEFT JOIN asistencia a ON t.id = a.trabajador_id
        WHERE a.hora_ingreso IS NOT NULL
    '''
    
    params = []
    if fecha_inicio:
        query += ' AND a.fecha >= ?'
        params.append(fecha_inicio)
    if fecha_fin:
        query += ' AND a.fecha <= ?'
        params.append(fecha_fin)
    
    query += ' GROUP BY t.id ORDER BY t.nombre'
    
    cursor.execute(query, params)
    produccion = cursor.fetchall()
    conexion.close()
    
    return render_template('produccion.html',
                           usuario=session['usuario'],
                           produccion=produccion,
                           fecha_inicio=fecha_inicio,
                           fecha_fin=fecha_fin)

@app.route('/exportar')
def exportar():
    if 'usuario' not in session:
        return redirect(url_for('inicio'))
    return render_template('exportar.html', usuario=session['usuario'])

@app.route('/exportar/excel')
def exportar_excel():
    if 'usuario' not in session:
        return redirect(url_for('inicio'))

    import sqlite3
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from flask import send_file

    fecha_inicio = request.args.get('fecha_inicio', '')
    fecha_fin = request.args.get('fecha_fin', '')

    conexion = sqlite3.connect('asistencia.db')
    conexion.row_factory = sqlite3.Row
    cursor = conexion.cursor()
    cursor.execute('''
        SELECT t.nombre, t.dni, t.cargo, t.obra,
               a.fecha, a.hora_ingreso, a.hora_salida,
               a.ubicacion_ingreso, a.ubicacion_salida
        FROM trabajadores t
        LEFT JOIN asistencia a ON t.id = a.trabajador_id
        WHERE a.fecha BETWEEN ? AND ?
        ORDER BY a.fecha, t.nombre
    ''', (fecha_inicio, fecha_fin))
    datos = cursor.fetchall()
    conexion.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Asistencia"

    # NUEVO: agrega columnas de ubicación
    encabezados = ['Nombre', 'DNI', 'Cargo', 'Obra', 'Fecha', 
                   'Hora Ingreso', 'Ubicacion Ingreso', 
                   'Hora Salida', 'Ubicacion Salida']
    rojo = PatternFill("solid", fgColor="C0392B")
    for col, titulo in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = rojo
        celda.alignment = Alignment(horizontal="center")

    for fila, row in enumerate(datos, 2):
        ws.cell(row=fila, column=1, value=row['nombre'])
        ws.cell(row=fila, column=2, value=row['dni'])
        ws.cell(row=fila, column=3, value=row['cargo'])
        ws.cell(row=fila, column=4, value=row['obra'])
        ws.cell(row=fila, column=5, value=row['fecha'])
        ws.cell(row=fila, column=6, value=row['hora_ingreso'] or 'Sin registro')
        ws.cell(row=fila, column=7, value=row['ubicacion_ingreso'] or 'Sin registro')
        ws.cell(row=fila, column=8, value=row['hora_salida'] or 'Sin registro')
        ws.cell(row=fila, column=9, value=row['ubicacion_salida'] or 'Sin registro')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'reporte_{fecha_inicio}_{fecha_fin}.xlsx')

@app.route('/exportar/pdf')
def exportar_pdf():
    if 'usuario' not in session:
        return redirect(url_for('inicio'))

    import sqlite3
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from io import BytesIO
    from flask import send_file

    fecha_inicio = request.args.get('fecha_inicio', '')
    fecha_fin = request.args.get('fecha_fin', '')

    conexion = sqlite3.connect('asistencia.db')
    conexion.row_factory = sqlite3.Row
    cursor = conexion.cursor()
    cursor.execute('''
        SELECT t.nombre, t.dni, t.cargo, t.obra,
               a.fecha, a.hora_ingreso, a.hora_salida,
               a.ubicacion_ingreso, a.ubicacion_salida
        FROM trabajadores t
        LEFT JOIN asistencia a ON t.id = a.trabajador_id
        WHERE a.fecha BETWEEN ? AND ?
        ORDER BY a.fecha, t.nombre
    ''', (fecha_inicio, fecha_fin))
    datos = cursor.fetchall()
    conexion.close()

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=20, rightMargin=20)
    styles = getSampleStyleSheet()
    elementos = []

    titulo = Paragraph(f"<b>Reporte de Asistencia</b> — {fecha_inicio} al {fecha_fin}", styles['Title'])
    elementos.append(titulo)
    elementos.append(Spacer(1, 16))

    # NUEVO: agrega columnas de ubicación
    encabezados = [['Nombre', 'DNI', 'Cargo', 'Obra', 'Fecha',
                    'H. Ingreso', 'Ubic. Ingreso', 'H. Salida', 'Ubic. Salida']]
    filas = encabezados + [
        [row['nombre'], row['dni'], row['cargo'], row['obra'], row['fecha'],
         row['hora_ingreso'] or '—', row['ubicacion_ingreso'] or '—',
         row['hora_salida'] or '—', row['ubicacion_salida'] or '—']
        for row in datos
    ]

    tabla = Table(filas, colWidths=[90, 60, 70, 70, 60, 55, 100, 55, 100])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#C0392B')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,0), 9),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#fdf0f0')]),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#e0e0e0')),
        ('FONTSIZE',   (0,1), (-1,-1), 8),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('ROWHEIGHT',  (0,0), (-1,-1), 20),
    ]))
    elementos.append(tabla)

    doc.build(elementos)
    output.seek(0)

    return send_file(output,
                     mimetype='application/pdf',
                     as_attachment=True,
                     download_name=f'reporte_{fecha_inicio}_{fecha_fin}.pdf')

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
    J